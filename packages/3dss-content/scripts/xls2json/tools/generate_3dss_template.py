# generate_3dss_template.py
import json
from pathlib import Path
from typing import Any, Dict, List, Tuple, Optional
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.worksheet.datavalidation import DataValidation

REQUIRED_FILL = PatternFill("solid", fgColor="FFF2CC")  # required: light yellow
DEFAULT_FILL  = PatternFill("solid", fgColor="E6E6E6")  # default: light gray

def load_schema(path: Path) -> Dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))

def flatten_properties(
    schema: Dict[str, Any],
    base_required: Optional[set] = None,
    prefix: str = ""
) -> List[Dict[str, Any]]:
    """
    schema: object schema with 'properties'
    returns list of column specs: {path,type,required,default,enum,description}
    """
    cols: List[Dict[str, Any]] = []
    required = set(schema.get("required", []))
    if base_required:
        required |= base_required

    props = schema.get("properties", {})
    for k, v in props.items():
        path = f"{prefix}.{k}" if prefix else k

        t = v.get("type")
        enum = v.get("enum")
        default = v.get("default")
        desc = v.get("description", "")

        # nested object => recurse
        if t == "object" and "properties" in v:
            cols.extend(flatten_properties(v, None, path))
            continue

        # arrays / free-form maps => keep as json cell
        if t == "array" or (t == "object" and v.get("additionalProperties")):
            cols.append({
                "path": path + "_json",
                "type": "json",
                "required": path in required,
                "default": default,
                "enum": None,
                "description": (desc + " (JSON)").strip(),
                "raw": v,
            })
            continue

        cols.append({
            "path": path,
            "type": t or "any",
            "required": path in required,
            "default": default,
            "enum": enum,
            "description": desc,
            "raw": v,
        })
    return cols

def write_sheet(ws, cols: List[Dict[str, Any]]):
    # Row 1: header (paths)
    # Row 2: type
    # Row 3: description
    ws.freeze_panes = "A4"

    for c, col in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=c, value=col["path"])
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

        ws.cell(row=2, column=c, value=col["type"])
        ws.cell(row=3, column=c, value=col["description"])

        if col["required"]:
            for r in (1,2,3):
                ws.cell(row=r, column=c).fill = REQUIRED_FILL

        if col["default"] is not None:
            ws.cell(row=2, column=c, value=f'{col["type"]} (default={col["default"]})')
            for r in (1,2,3):
                ws.cell(row=r, column=c).fill = DEFAULT_FILL if not col["required"] else REQUIRED_FILL

        # enum dropdown
        if col["enum"]:
            dv = DataValidation(type="list", formula1=f'"{",".join(map(str,col["enum"]))}"', allow_blank=not col["required"])
            ws.add_data_validation(dv)
            dv.add(ws.cell(row=4, column=c))  # applies starting row (copy down in Excel)

    ws.auto_filter.ref = ws.dimensions

def main(schema_path: str, out_xlsx: str):
    schema = load_schema(Path(schema_path))

    # ↓ここはスキーマ構造に合わせて「points の定義」「lines の定義」を指す場所を調整する必要あり
    points_schema = schema["$defs"]["point"]   # 例
    lines_schema  = schema["$defs"]["line"]    # 例

    wb = Workbook()
    wb.remove(wb.active)

    ws_p = wb.create_sheet("points")
    cols_p = flatten_properties(points_schema)
    write_sheet(ws_p, cols_p)

    ws_l = wb.create_sheet("lines")
    cols_l = flatten_properties(lines_schema)
    write_sheet(ws_l, cols_l)

    wb.save(out_xlsx)

if __name__ == "__main__":
    main("3DSS.schema.json", "3DSS_template.xlsx")
