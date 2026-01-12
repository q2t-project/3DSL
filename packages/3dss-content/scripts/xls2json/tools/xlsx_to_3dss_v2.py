#!/usr/bin/env python3
# xlsx_to_3dss_v2.py
# Convert a 2-sheet Excel workbook ("points", "lines") plus optional "document_meta" into 3DSS.json.
#
# Expected layout per element sheet:
#   Row 1: column keys (JSON paths)
#   Row 2: type info (e.g., "string", "number (default=...)", "json", etc.)
#   Row 3: description / notes
#   Row 4..: data rows (1 element per row)
#
# Optional sheet "document_meta":
#   Row 1: ["key","value"]
#   Row 2..: key/value rows. Values that look like JSON are parsed (objects/arrays/numbers/bools/null).
#
# Usage:
#   python xlsx_to_3dss_v2.py --xlsx INPUT.xlsx --schema 3DSS.schema.json --out OUT.3dss.json
#
import json
import re
import uuid
import datetime
import argparse
from pathlib import Path
from typing import Any, Dict, List, Tuple, Optional

from openpyxl import load_workbook

try:
    import jsonschema
except Exception:
    jsonschema = None

ARRAY_IDX_RE = re.compile(r"^(?P<name>[^\[\]]+)(?:\[(?P<idx>\d+)\])?$")

def _parse_steps(key: str) -> List[Tuple[Optional[str], Optional[int]]]:
    steps: List[Tuple[Optional[str], Optional[int]]] = []
    for seg in key.split("."):
        m = ARRAY_IDX_RE.match(seg)
        if not m:
            steps.append((seg, None))
            continue
        name = m.group("name")
        idx = m.group("idx")
        steps.append((name, int(idx)) if idx is not None else (name, None))
    return steps

def _ensure_list_len(lst: List[Any], n: int) -> None:
    if len(lst) < n:
        lst.extend([None] * (n - len(lst)))

def _set_path(obj: Dict[str, Any], key: str, value: Any) -> None:
    steps = _parse_steps(key)
    cur: Any = obj
    for i, (name, idx) in enumerate(steps):
        is_last = (i == len(steps) - 1)
        if name is None:
            return

        if idx is None:
            if is_last:
                cur[name] = value
                return
            if name not in cur or not isinstance(cur[name], (dict, list)):
                nxt_name, nxt_idx = steps[i + 1]
                cur[name] = [] if nxt_idx is not None else {}
            cur = cur[name]
        else:
            if name not in cur or not isinstance(cur[name], list):
                cur[name] = []
            lst = cur[name]
            _ensure_list_len(lst, idx + 1)
            if is_last:
                lst[idx] = value
                return
            if lst[idx] is None or not isinstance(lst[idx], (dict, list)):
                nxt_name, nxt_idx = steps[i + 1]
                lst[idx] = [] if nxt_idx is not None else {}
            cur = lst[idx]

def _trim(obj: Any) -> Any:
    if isinstance(obj, dict):
        out = {}
        for k, v in obj.items():
            v2 = _trim(v)
            if v2 is None:
                continue
            if isinstance(v2, dict) and not v2:
                continue
            if isinstance(v2, list) and len(v2) == 0:
                continue
            out[k] = v2
        return out
    if isinstance(obj, list):
        out = [_trim(v) for v in obj]
        while out and out[-1] is None:
            out.pop()
        if all(v is None for v in out):
            return []
        return out
    return obj

def _base_type(type_cell: Any) -> str:
    if type_cell is None:
        return "any"
    s = str(type_cell).strip()
    if not s:
        return "any"
    return s.split()[0].lower()

def _coerce(val: Any, base_type: str, key: str) -> Any:
    if val is None:
        return None

    if isinstance(val, (datetime.datetime, datetime.date)):
        return val.isoformat()

    if base_type == "json" or key.endswith("_json"):
        if isinstance(val, (dict, list)):
            return val
        s = str(val).strip()
        if not s:
            return None
        try:
            return json.loads(s)
        except Exception:
            return s

    if base_type == "boolean":
        if isinstance(val, bool):
            return val
        s = str(val).strip().lower()
        if s in ("true", "1", "yes", "y", "on"):
            return True
        if s in ("false", "0", "no", "n", "off"):
            return False
        return None

    if base_type == "integer":
        if isinstance(val, int) and not isinstance(val, bool):
            return val
        if isinstance(val, float):
            return int(val)
        s = str(val).strip()
        if not s:
            return None
        try:
            return int(float(s))
        except Exception:
            return None

    if base_type == "number":
        if isinstance(val, (int, float)) and not isinstance(val, bool):
            return float(val)
        s = str(val).strip()
        if not s:
            return None
        try:
            return float(s)
        except Exception:
            return None

    if isinstance(val, str):
        s = val.strip()
        return s if s != "" else None
    return val

def _read_sheet(wb, sheet_name: str) -> List[Dict[str, Any]]:
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    keys = [c.value for c in ws[1]]
    types = [c.value for c in ws[2]]

    col_map = []
    for idx, key in enumerate(keys):
        if key is None:
            continue
        key_s = str(key).strip()
        if not key_s:
            continue
        col_map.append((idx + 1, key_s, _base_type(types[idx] if idx < len(types) else None)))

    rows: List[Dict[str, Any]] = []
    for r in range(4, ws.max_row + 1):
        obj: Dict[str, Any] = {}
        any_value = False
        for col_idx, key_s, base_t in col_map:
            cell = ws.cell(row=r, column=col_idx)
            v = _coerce(cell.value, base_t, key_s)
            if v is None:
                continue
            any_value = True
            _set_path(obj, key_s, v)

        if not any_value:
            continue

        obj = _trim(obj)
        if isinstance(obj, dict):
            meta = obj.get("meta")
            if isinstance(meta, dict) and not meta.get("uuid"):
                meta["uuid"] = str(uuid.uuid4())
            elif meta is None:
                obj["meta"] = {"uuid": str(uuid.uuid4())}
        rows.append(obj)

    return rows

def _parse_meta_value(v: Any) -> Any:
    if v is None:
        return None
    if isinstance(v, (int, float, bool, datetime.datetime, datetime.date)):
        return v if not isinstance(v, (datetime.datetime, datetime.date)) else v.isoformat()
    if not isinstance(v, str):
        return v
    s = v.strip()
    if s == "":
        return None
    # Try JSON parse for objects/arrays/number/bool/null
    try:
        return json.loads(s)
    except Exception:
        return s

def _read_document_meta(wb) -> Optional[Dict[str, Any]]:
    if "document_meta" not in wb.sheetnames:
        return None
    ws = wb["document_meta"]
    # header in row 1: key, value
    meta: Dict[str, Any] = {}
    for r in range(2, ws.max_row + 1):
        k = ws.cell(row=r, column=1).value
        if k is None:
            continue
        ks = str(k).strip()
        if not ks:
            continue
        v = ws.cell(row=r, column=2).value
        meta[ks] = _parse_meta_value(v)
    return meta if meta else None

def _default_document_meta(schema: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
    schema_uri = "https://q2t-project.github.io/3dsl/schemas/3DSS.schema.json#v1.1.3"
    if schema and isinstance(schema, dict):
        sid = schema.get("$id") or ""
        anch = schema.get("$anchor") or ""
        if sid:
            base = sid[:-1] if sid.endswith("#") else sid
            if anch:
                schema_uri = f"{base}#{anch}"
            else:
                schema_uri = base

    return {
        "document_title": "Untitled",
        "document_uuid": str(uuid.uuid4()),
        "schema_uri": schema_uri,
        "author": "unknown",
        "version": "1.0.0",
        "updated_at": datetime.datetime.utcnow().replace(microsecond=0).isoformat() + "Z",
    }

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", required=True, help="Input .xlsx (sheets: points, lines)")
    ap.add_argument("--out", required=True, help="Output .json path")
    ap.add_argument("--schema", default=None, help="Optional 3DSS.schema.json to validate output")
    ap.add_argument("--no-validate", action="store_true", help="Skip schema validation even if --schema is given")
    args = ap.parse_args()

    schema = None
    if args.schema:
        schema = json.loads(Path(args.schema).read_text(encoding="utf-8"))

    wb = load_workbook(args.xlsx, data_only=True)

    points = _read_sheet(wb, "points")
    lines = _read_sheet(wb, "lines")

    document_meta = _read_document_meta(wb) or _default_document_meta(schema)

    # If user left a placeholder in document_uuid, auto-fill
    if isinstance(document_meta, dict):
        du = document_meta.get("document_uuid")
        if not du or (isinstance(du, str) and "PUT_UUID" in du):
            document_meta["document_uuid"] = str(uuid.uuid4())

    doc = {
        "document_meta": document_meta,
        "points": points,
        "lines": lines,
    }

    Path(args.out).write_text(json.dumps(doc, ensure_ascii=False, indent=2), encoding="utf-8")

    if args.schema and not args.no_validate and jsonschema is not None:
        try:
            jsonschema.validate(instance=doc, schema=schema)
        except Exception as e:
            raise SystemExit(f"[validate] FAILED: {e}")
        print("[validate] OK")

    print(f"[write] {args.out} (points={len(points)} lines={len(lines)})")

if __name__ == "__main__":
    main()
