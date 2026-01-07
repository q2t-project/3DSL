#!/usr/bin/env python3
# json_to_xlsx.py
# Convert 3DSS.json -> Excel workbook using an existing template workbook (preferred),
# preserving styles/validations. Also writes a "document_meta" sheet for round-tripping.
#
# Usage:
#   python json_to_xlsx.py --json INPUT.3dss.json --template 3DSS_points_lines_template.xlsx --out OUT.xlsx
#
import json
import re
import argparse
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

ARRAY_IDX_RE = re.compile(r"^(?P<name>[^\[\]]+)(?:\[(?P<idx>\d+)\])?$")

def _parse_steps(key: str) -> List[Tuple[str, Optional[int]]]:
    steps: List[Tuple[str, Optional[int]]] = []
    for seg in key.split("."):
        m = ARRAY_IDX_RE.match(seg)
        if not m:
            steps.append((seg, None))
            continue
        name = m.group("name")
        idx = m.group("idx")
        steps.append((name, int(idx)) if idx is not None else (name, None))
    return steps

def _get_path(obj: Any, key: str) -> Any:
    steps = _parse_steps(key)
    cur: Any = obj
    for name, idx in steps:
        if not isinstance(cur, dict):
            return None
        if name not in cur:
            return None
        cur = cur[name]
        if idx is not None:
            if not isinstance(cur, list):
                return None
            if idx < 0 or idx >= len(cur):
                return None
            cur = cur[idx]
    return cur

def _to_cell_value(v: Any, key: str) -> Any:
    if v is None:
        return None
    # JSON columns: store as JSON string for safe round-trip
    if key.endswith("_json"):
        try:
            return json.dumps(v, ensure_ascii=False)
        except Exception:
            return str(v)

    # If a dict/list accidentally appears in scalar column, stringify JSON
    if isinstance(v, (dict, list)):
        try:
            return json.dumps(v, ensure_ascii=False)
        except Exception:
            return str(v)

    return v

def _clear_data_rows(ws, start_row: int = 4, max_rows: int = 5000):
    # Keep header rows; clear existing data region
    for r in range(start_row, min(ws.max_row, max_rows) + 1):
        for c in range(1, ws.max_column + 1):
            ws.cell(row=r, column=c).value = None

def _write_elements(ws, elements: List[Dict[str, Any]], start_row: int = 4, max_rows: int = 5000):
    # Column keys are in row 1
    keys = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    col_keys: List[Tuple[int, str]] = []
    for c, k in enumerate(keys, start=1):
        if k is None:
            continue
        ks = str(k).strip()
        if not ks:
            continue
        col_keys.append((c, ks))

    _clear_data_rows(ws, start_row=start_row, max_rows=max_rows)

    r = start_row
    for el in elements:
        if r > max_rows:
            raise SystemExit(f"Too many rows; exceeded max_rows={max_rows}")
        for c, ks in col_keys:
            v = _get_path(el, ks)
            ws.cell(row=r, column=c).value = _to_cell_value(v, ks)
        r += 1

def _write_document_meta(wb, document_meta: Dict[str, Any]):
    # Create or clear "document_meta" sheet
    if "document_meta" in wb.sheetnames:
        ws = wb["document_meta"]
        # clear
        for r in range(1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                ws.cell(row=r, column=c).value = None
    else:
        ws = wb.create_sheet("document_meta")

    ws.cell(row=1, column=1, value="key")
    ws.cell(row=1, column=2, value="value")

    r = 2
    for k in sorted(document_meta.keys()):
        v = document_meta[k]
        ws.cell(row=r, column=1, value=k)
        if isinstance(v, (dict, list)):
            ws.cell(row=r, column=2, value=json.dumps(v, ensure_ascii=False))
        else:
            ws.cell(row=r, column=2, value=v)
        r += 1

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--json", required=True, help="Input 3DSS.json")
    ap.add_argument("--template", required=True, help="Template .xlsx (must contain points/lines sheets)")
    ap.add_argument("--out", required=True, help="Output .xlsx")
    ap.add_argument("--max-rows", type=int, default=5000, help="Max rows per sheet to write")
    args = ap.parse_args()

    doc = json.loads(Path(args.json).read_text(encoding="utf-8"))
    points = doc.get("points", []) or []
    lines = doc.get("lines", []) or []
    document_meta = doc.get("document_meta", {}) or {}

    wb = load_workbook(args.template)
    if "points" in wb.sheetnames:
        _write_elements(wb["points"], points, max_rows=args.max_rows)
    else:
        raise SystemExit("Template missing sheet: points")
    if "lines" in wb.sheetnames:
        _write_elements(wb["lines"], lines, max_rows=args.max_rows)
    else:
        raise SystemExit("Template missing sheet: lines")

    _write_document_meta(wb, document_meta)

    Path(args.out).parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.out)
    print(f"[write] {args.out} (points={len(points)} lines={len(lines)})")

if __name__ == "__main__":
    main()
